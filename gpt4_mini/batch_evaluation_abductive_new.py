#!/usr/bin/env python3
"""
Standalone Batch Evaluation Script for Abductive Predictions
Reads prediction CSVs, downloads ground truth from HuggingFace, computes metrics, saves xlsx.
"""

import os
import sys
import re
import pandas as pd
import numpy as np
from datetime import datetime
from datasets import load_dataset
import requests
from io import StringIO

# ============================================================================
# CONFIGURATION — edit these
# ============================================================================
PREDICTIONS_DIR = "abductive_gpt-4o-mini_predictions_csv"
MODEL_NAME      = "gpt-4o-mini"
REASONING_TYPE  = "abductive"
HF_TOKEN = os.getenv("HF_TOKEN", "")   # hard-code your HuggingFace token here
OUTPUT_XLSX     = f"batch_evaluation_results_{MODEL_NAME}_{REASONING_TYPE}.xlsx"

# ============================================================================
# DATASET CONFIG
# ============================================================================
DATASETS = [
    "ASAP-AES",
    "ASAP2",
    "ASAP-SAS",
    "ASAP_plus_plus",
    "BEEtlE_2way",
    "BEEtlE_3way",
    "SciEntSBank_2way",
    "SciEntSBank_3way",
    "CSEE",
    "Mohlar",
    "Ielts_Writing_Dataset",
    "Ielts_Writing_Task_2_Dataset",
    "persuade_2",
    "Regrading_Dataset_J2C",
    "OS_Dataset_q1",
    "OS_Dataset_q2",
    "OS_Dataset_q3",
    "OS_Dataset_q4",
    "OS_Dataset_q5",
    "Rice_Chem_Q1",
    "Rice_Chem_Q2",
    "Rice_Chem_Q3",
    "Rice_Chem_Q4",
]

# Column configs: id, prediction column in CSV, ground truth column in HF
DATASET_CONFIG = {
    "ASAP-AES":                  {"id": "essay_id",       "pred_col": "domain1_score",       "gt_col": "domain1_score",       "essay_set_col": "essay_set"},
    "ASAP2":                     {"id": "essay_id",       "pred_col": "score",                "gt_col": "score",                "essay_set_col": None},
    "ASAP-SAS":                  {"id": "Id",             "pred_col": "Score1",               "gt_col": "Score1",               "essay_set_col": None},
    "ASAP_plus_plus":            {"id": "essay_id",       "pred_col": "overall_score",        "gt_col": "overall_score",        "essay_set_col": "essay_set"},
    "BEEtlE_2way":               {"id": "ID",             "pred_col": "label",                "gt_col": "label",                "essay_set_col": None},
    "BEEtlE_3way":               {"id": "ID",             "pred_col": "label",                "gt_col": "label",                "essay_set_col": None},
    "SciEntSBank_2way":          {"id": "ID",             "pred_col": "label",                "gt_col": "label",                "essay_set_col": None},
    "SciEntSBank_3way":          {"id": "ID",             "pred_col": "label",                "gt_col": "label",                "essay_set_col": None},
    "CSEE":                      {"id": "index",          "pred_col": "overall_score",        "gt_col": "overall_score",        "essay_set_col": None},
    "Mohlar":                    {"id": "ID",             "pred_col": "grade",                "gt_col": "grade",                "essay_set_col": None},
    "Ielts_Writing_Dataset":     {"id": "ID",             "pred_col": "Overall_Score",        "gt_col": "Overall_Score",        "essay_set_col": None},
    "Ielts_Writing_Task_2_Dataset": {"id": "ID",          "pred_col": "band_score",           "gt_col": "band_score",           "essay_set_col": None},
    "persuade_2":                {"id": "essay_id_comp",  "pred_col": "holistic_essay_score", "gt_col": "holistic_essay_score", "essay_set_col": None},
    "Regrading_Dataset_J2C":     {"id": "ID",             "pred_col": "grade",                "gt_col": "grade",                "essay_set_col": None},
    "OS_Dataset_q1":             {"id": "ID",             "pred_col": "score_1",              "gt_col": "score_1",              "essay_set_col": None},
    "OS_Dataset_q2":             {"id": "ID",             "pred_col": "score_1",              "gt_col": "score_1",              "essay_set_col": None},
    "OS_Dataset_q3":             {"id": "ID",             "pred_col": "score_1",              "gt_col": "score_1",              "essay_set_col": None},
    "OS_Dataset_q4":             {"id": "ID",             "pred_col": "score_1",              "gt_col": "score_1",              "essay_set_col": None},
    "OS_Dataset_q5":             {"id": "ID",             "pred_col": "score_1",              "gt_col": "score_1",              "essay_set_col": None},
    "Rice_Chem_Q1":              {"id": "sis_id",         "pred_col": "Score",                "gt_col": "Score",                "essay_set_col": None},
    "Rice_Chem_Q2":              {"id": "sis_id",         "pred_col": "Score",                "gt_col": "Score",                "essay_set_col": None},
    "Rice_Chem_Q3":              {"id": "sis_id",         "pred_col": "Score",                "gt_col": "Score",                "essay_set_col": None},
    "Rice_Chem_Q4":              {"id": "sis_id",         "pred_col": "Score",                "gt_col": "Score",                "essay_set_col": None},
}

SCORE_RANGES = {
    "ASAP-AES":    {1:(2,12), 2:(1,6), 3:(0,3), 4:(0,3), 5:(0,4), 6:(0,4), 7:(0,30), 8:(0,60)},
    "ASAP_plus_plus": {1:(2,12), 2:(1,6), 3:(0,3), 4:(0,3), 5:(0,4), 6:(0,4), 7:(0,30), 8:(0,60)},
    "ASAP2":                     (0, 3),
    "ASAP-SAS":                  (0, 3),
    "BEEtlE_2way":               (0, 1),
    "BEEtlE_3way":               (0, 2),
    "SciEntSBank_2way":          (0, 1),
    "SciEntSBank_3way":          (0, 2),
    "CSEE":                      (0, 16),
    "Mohlar":                    (0, 5),
    "Ielts_Writing_Dataset":     (1, 9),
    "Ielts_Writing_Task_2_Dataset": (1, 9),
    "persuade_2":                (1, 6),
    "Regrading_Dataset_J2C":     (0, 8),
    "OS_Dataset_q1":             (0, 19),
    "OS_Dataset_q2":             (0, 16),
    "OS_Dataset_q3":             (0, 15),
    "OS_Dataset_q4":             (0, 16),
    "OS_Dataset_q5":             (0, 27),
    "Rice_Chem_Q1":              (0, 8),
    "Rice_Chem_Q2":              (0, 8),
    "Rice_Chem_Q3":              (0, 9),
    "Rice_Chem_Q4":              (0, 8),
}

CLASSIFICATION_DATASETS = ["BEEtlE_2way", "BEEtlE_3way", "SciEntSBank_2way", "SciEntSBank_3way"]

# ============================================================================
# GROUND TRUTH DOWNLOADER
# ============================================================================
def download_ground_truth(dataset_name: str, hf_token: str) -> pd.DataFrame:
    headers = {"Authorization": f"Bearer {hf_token}"} if hf_token else {}

    print(f"  Downloading ground truth for {dataset_name}...")

    try:
        if dataset_name in ["BEEtlE_2way", "BEEtlE_3way"]:
            suffix = "2way" if "2way" in dataset_name else "3way"
            url = f"https://huggingface.co/datasets/nlpatunt/BEEtlE/resolve/main/test_{suffix}.csv"
            r = requests.get(url, headers=headers, timeout=60)
            r.raise_for_status()
            return pd.read_csv(StringIO(r.text))

        elif dataset_name in ["SciEntSBank_2way", "SciEntSBank_3way"]:
            suffix = "2way" if "2way" in dataset_name else "3way"
            url = f"https://huggingface.co/datasets/nlpatunt/SciEntSBank/resolve/main/test_{suffix}.csv"
            r = requests.get(url, headers=headers, timeout=60)
            r.raise_for_status()
            return pd.read_csv(StringIO(r.text))

        elif dataset_name == "ASAP-SAS":
            url = "https://huggingface.co/datasets/nlpatunt/ASAP-SAS/resolve/main/test.csv"
            r = requests.get(url, headers=headers, timeout=60)
            r.raise_for_status()
            return pd.read_csv(StringIO(r.text))

        elif dataset_name in ["Rice_Chem_Q1", "Rice_Chem_Q2", "Rice_Chem_Q3", "Rice_Chem_Q4"]:
            q_num = dataset_name.split("_")[-1]
            ds = load_dataset("nlpatunt/Rice_Chem", data_files=f"{q_num}/test.csv", token=hf_token)["train"]
            return ds.to_pandas()

        elif dataset_name.startswith("OS_Dataset_q"):
            q_num = dataset_name.split("_q")[-1]
            ds = load_dataset("nlpatunt/OS_Dataset", data_files=f"q{q_num}/test.csv", token=hf_token)["train"]
            return ds.to_pandas()

        elif dataset_name == "Mohlar":
            url = "https://huggingface.co/datasets/nlpatunt/Mohlar/resolve/main/test.csv"
            r = requests.get(url, headers=headers, timeout=60)
            r.raise_for_status()
            return pd.read_csv(StringIO(r.text), usecols=lambda c: not c.startswith("Unnamed"))

        elif dataset_name == "persuade_2":
            ds = load_dataset("nlpatunt/persuade_2", split="test", token=hf_token)
            return ds.to_pandas()

        else:
            ds = load_dataset(f"nlpatunt/{dataset_name}", split="test", token=hf_token, trust_remote_code=True)
            return ds.to_pandas()

    except Exception as e:
        print(f"  Failed to download ground truth for {dataset_name}: {e}")
        return None

# ============================================================================
# PREDICTION CSV LOADER
# ============================================================================
def load_predictions(dataset_name: str) -> pd.DataFrame:
    csv_path = os.path.join(PREDICTIONS_DIR, f"{MODEL_NAME}_{dataset_name}_{REASONING_TYPE}.csv")
    if not os.path.exists(csv_path):
        print(f"  ✗ Prediction file not found: {csv_path}")
        return None
    df = pd.read_csv(csv_path)
    print(f"  ✓ Loaded predictions: {len(df)} rows from {csv_path}")
    return df

# ============================================================================
# METRICS CALCULATOR
# ============================================================================
def calculate_metrics(y_true, y_pred, dataset_name: str, essay_sets=None):
    from sklearn.metrics import (
        mean_absolute_error, mean_squared_error, f1_score,
        precision_score, recall_score, cohen_kappa_score, accuracy_score
    )
    from scipy.stats import pearsonr

    is_classification = dataset_name in CLASSIFICATION_DATASETS

    if is_classification:
        if "3way" in dataset_name:
            label_map = {"correct": 2, "incorrect": 0, "contradictory": 1}
        else:
            label_map = {"correct": 1, "incorrect": 0}

        y_true_num = np.array([label_map.get(str(v).strip().lower(), -1) for v in y_true])
        y_pred_num = np.array([label_map.get(str(v).strip().lower(), -1) for v in y_pred])

        # Filter invalid
        valid = (y_true_num >= 0) & (y_pred_num >= 0)
        y_true_num = y_true_num[valid]
        y_pred_num = y_pred_num[valid]

        if len(y_true_num) == 0:
            return None

        accuracy = accuracy_score(y_true_num, y_pred_num)
        try:
            qwk = cohen_kappa_score(y_true_num, y_pred_num, weights="quadratic")
        except:
            qwk = 0.0
        f1        = f1_score(y_true_num, y_pred_num, average="weighted", zero_division=0)
        precision = precision_score(y_true_num, y_pred_num, average="weighted", zero_division=0)
        recall    = recall_score(y_true_num, y_pred_num, average="weighted", zero_division=0)
        pearson   = accuracy  # Pearson not meaningful for classification
        mae       = 1.0 - accuracy
        rmse      = np.sqrt((1.0 - accuracy) ** 2)
        mae_pct   = None

    else:
        y_true = pd.to_numeric(y_true, errors='coerce')
        y_pred = pd.to_numeric(y_pred, errors='coerce')
        y_true = np.array(y_true, dtype=np.float64)
        y_pred = np.array(y_pred, dtype=np.float64)
        valid  = ~(np.isnan(y_true) | np.isnan(y_pred))
        y_true = y_true[valid]
        y_pred = y_pred[valid]

        if len(y_true) < 2:
            return None

        pearson, _ = pearsonr(y_true, y_pred)
        mae        = mean_absolute_error(y_true, y_pred)
        mse        = mean_squared_error(y_true, y_pred)
        rmse       = np.sqrt(mse)

        y_true_int = np.round(y_true).astype(int)
        y_pred_int = np.round(y_pred).astype(int)

        accuracy  = accuracy_score(y_true_int, y_pred_int)
        try:
            qwk = cohen_kappa_score(y_true_int, y_pred_int, weights="quadratic")
        except:
            qwk = 0.0
        f1        = f1_score(y_true_int, y_pred_int, average="weighted", zero_division=0)
        precision = precision_score(y_true_int, y_pred_int, average="weighted", zero_division=0)
        recall    = recall_score(y_true_int, y_pred_int, average="weighted", zero_division=0)

        # MAE %
        if dataset_name in ["ASAP-AES", "ASAP_plus_plus"] and essay_sets is not None:
            mae_pcts = []
            for es in np.unique(essay_sets[valid]):
                mask = essay_sets[valid] == es
                if mask.sum() == 0:
                    continue
                r = SCORE_RANGES[dataset_name].get(int(es), (0, 12))
                range_size = r[1] - r[0]
                if range_size > 0:
                    mae_pcts.append(mean_absolute_error(y_true[mask], y_pred[mask]) / range_size * 100)
            mae_pct = round(np.mean(mae_pcts), 2) if mae_pcts else None
        else:
            r = SCORE_RANGES.get(dataset_name, (0, 100))
            range_size = r[1] - r[0] if isinstance(r, tuple) else 100
            mae_pct = round(mae / range_size * 100, 2) if range_size > 0 else None

    return {
        "QWK":       round(float(qwk), 6),
        "Pearson":   round(float(pearson), 6),
        "F1":        round(float(f1), 6),
        "Precision": round(float(precision), 6),
        "Recall":    round(float(recall), 6),
        "Accuracy":  round(float(accuracy), 6),
        "MAE":       round(float(mae), 6),
        "RMSE":      round(float(rmse), 6),
        "MAE_pct":   mae_pct,
        "n":         len(y_true) if not is_classification else int(valid.sum()),
    }

# ============================================================================
# MAIN EVALUATION LOOP
# ============================================================================
def run_evaluation():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print("=" * 70)
    print(f"BATCH EVALUATION — {MODEL_NAME} — {REASONING_TYPE}")
    print("=" * 70)

    all_results = []

    for dataset_name in DATASETS:
        print(f"\n{'='*60}")
        print(f"Dataset: {dataset_name}")

        cfg = DATASET_CONFIG[dataset_name]
        id_col    = cfg["id"]
        pred_col  = cfg["pred_col"]
        gt_col    = cfg["gt_col"]
        es_col    = cfg.get("essay_set_col")

        # Load predictions
        pred_df = load_predictions(dataset_name)
        if pred_df is None:
            all_results.append({"Dataset": dataset_name, "Error": "Prediction file not found"})
            continue

        # Download ground truth
        gt_df = download_ground_truth(dataset_name, HF_TOKEN)
        if gt_df is None:
            all_results.append({"Dataset": dataset_name, "Error": "Ground truth download failed"})
            continue

        print(f"  GT columns: {list(gt_df.columns)}")
        print(f"  Pred columns: {list(pred_df.columns)}")

        # Ensure ID cols are strings
        pred_df[id_col] = pred_df[id_col].astype(str)
        gt_df[id_col]   = gt_df[id_col].astype(str)

        # Merge
        gt_cols_to_merge = [id_col, gt_col]
        if es_col and es_col in gt_df.columns:
            gt_cols_to_merge.append(es_col)

        merged = pred_df.merge(
            gt_df[gt_cols_to_merge],
            on=id_col, how="inner", suffixes=("_pred", "_gt")
        )

        print(f"  Matched: {len(merged)} rows")

        if len(merged) == 0:
            all_results.append({"Dataset": dataset_name, "Error": "No matching IDs"})
            continue

        # Extract pred and gt values
        pred_col_merged = f"{pred_col}_pred" if f"{pred_col}_pred" in merged.columns else pred_col
        gt_col_merged   = f"{gt_col}_gt"   if f"{gt_col}_gt"   in merged.columns else gt_col

        y_pred = merged[pred_col_merged].values
        y_true = merged[gt_col_merged].values

        essay_sets = None
        if es_col and es_col in merged.columns:
            essay_sets = merged[es_col].values

        # Calculate metrics
        metrics = calculate_metrics(y_true, y_pred, dataset_name, essay_sets)

        if metrics is None:
            all_results.append({"Dataset": dataset_name, "Error": "Metrics calculation failed"})
            continue

        print(f"  QWK={metrics['QWK']:.4f} | Pearson={metrics['Pearson']:.4f} | F1={metrics['F1']:.4f} | MAE={metrics['MAE']:.4f} | n={metrics['n']}")

        all_results.append({
            "Dataset":        dataset_name,
            "Avg QWK":        metrics["QWK"],
            "Avg Pearson":    metrics["Pearson"],
            "Avg F1":         metrics["F1"],
            "Avg Precision":  metrics["Precision"],
            "Avg Recall":     metrics["Recall"],
            "Avg Accuracy":   metrics["Accuracy"],
            "Avg MAE":        metrics["MAE"],
            "Avg RMSE":       metrics["RMSE"],
            "MAE %":          metrics["MAE_pct"],
            "N":              metrics["n"],
        })

    # ========================================================================
    # COMPUTE AVERAGES
    # ========================================================================
    results_df = pd.DataFrame(all_results)

    # OS Dataset average (q1-q5)
    os_rows = results_df[results_df["Dataset"].str.startswith("OS_Dataset")]
    if len(os_rows) > 0:
        os_avg = os_rows[[c for c in os_rows.columns if c not in ["Dataset", "Error", "N"]]].mean(numeric_only=True)
        os_avg_row = {"Dataset": "📊 OS_Dataset Average (q1-q5)"}
        os_avg_row.update(os_avg.to_dict())
        results_df = pd.concat([results_df, pd.DataFrame([os_avg_row])], ignore_index=True)

    # Rice Chem average (Q1-Q4)
    rice_rows = results_df[results_df["Dataset"].str.startswith("Rice_Chem")]
    if len(rice_rows) > 0:
        rice_avg = rice_rows[[c for c in rice_rows.columns if c not in ["Dataset", "Error", "N"]]].mean(numeric_only=True)
        rice_avg_row = {"Dataset": "📊 Rice_Chem Average (Q1-Q4)"}
        rice_avg_row.update(rice_avg.to_dict())
        results_df = pd.concat([results_df, pd.DataFrame([rice_avg_row])], ignore_index=True)

    # Overall average (all datasets, no sub-averages)
    main_rows = results_df[~results_df["Dataset"].str.startswith("📊")]
    overall_avg = main_rows[[c for c in main_rows.columns if c not in ["Dataset", "Error", "N"]]].mean(numeric_only=True)
    overall_row = {"Dataset": "🏆 OVERALL AVERAGE (All Datasets)"}
    overall_row.update(overall_avg.to_dict())
    results_df = pd.concat([results_df, pd.DataFrame([overall_row])], ignore_index=True)

    # ========================================================================
    # SAVE XLSX
    # ========================================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Headers
    cols = ["Dataset", "Avg QWK", "Avg Pearson", "Avg F1", "Avg Precision",
            "Avg Recall", "Avg Accuracy", "Avg MAE", "Avg RMSE", "MAE %", "N"]

    header_fill   = PatternFill("solid", fgColor="1F4E79")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    avg_fill      = PatternFill("solid", fgColor="D6E4F0")
    overall_fill  = PatternFill("solid", fgColor="FFD700")
    overall_font  = Font(bold=True, size=11)

    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in results_df.iterrows():
        excel_row = row_idx + 2
        for col_idx, col_name in enumerate(cols, 1):
            val = row.get(col_name, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")

            dataset_val = str(row.get("Dataset", ""))
            if dataset_val.startswith("🏆"):
                cell.fill = overall_fill
                cell.font = overall_font
            elif dataset_val.startswith("📊"):
                cell.fill = avg_fill
                cell.font = Font(bold=True)

            # Round numeric values
            if isinstance(val, float) and col_name not in ["MAE %", "N"]:
                cell.value     = round(val, 6)
                cell.number_format = "0.000000"
            elif col_name == "MAE %" and isinstance(val, float):
                cell.number_format = "0.00"

    # Column widths
    ws.column_dimensions["A"].width = 35
    for col_idx in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    wb.save(OUTPUT_XLSX)
    print(f"\n✓ Saved: {OUTPUT_XLSX}")
    print(f"✓ {len(all_results)} datasets evaluated")

    # Print summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    for row in all_results:
        if "Error" in row:
            print(f"  ✗ {row['Dataset']}: {row['Error']}")
        else:
            print(f"  ✓ {row['Dataset']}: QWK={row['Avg QWK']:.4f} | F1={row['Avg F1']:.4f} | MAE={row['Avg MAE']:.4f}")

    return results_df

if __name__ == "__main__":
    run_evaluation()